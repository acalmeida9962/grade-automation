"""
Excel loading for the grade automation.

The teacher's workbook is messy in predictable ways, so parsing is defensive:

  * Row 1 may be blank — the header row is detected as the first row containing a
    cell whose text contains "nombre" (the student-name column).
  * The name-column header varies per sheet ("Nombre", "Nombre - Filosofía", ...).
  * Grade cells may hold: a number, "-" (means 1.5), blank (no grade), a DATE
    (Excel misparsed an entry like "4.5" into 2026-05-04 — decoded back as
    day + month/10), or free text (reported, never guessed).

This module only PARSES and CONVERTS values; matching student/column names to the
web platform (and the normalization used for that) lives in poc.py.
"""

from __future__ import annotations

import datetime
import math
import re
import unicodedata
from dataclasses import dataclass, field
from typing import Any

import openpyxl

DASH_GRADE = 1.5  # "-" in a grade cell means 1.5.


def _norm(s: Any) -> str:
    if s is None:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(s))
    no_accents = "".join(c for c in nfkd if not unicodedata.combining(c))
    return " ".join(no_accents.upper().split())


@dataclass
class Student:
    raw_name: str
    cells: dict[str, Any]  # grade header (original text) -> raw cell value


@dataclass
class ParsedSheet:
    sheet: str
    header_row: int
    name_header: str
    grade_headers: list[str]
    students: list[Student] = field(default_factory=list)


def list_sheets(path: str) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _find_header_row(ws, scan: int = 15) -> tuple[int, int]:
    """Return (header_row_index, name_col_index) — 1-based — else raise."""
    for r in range(1, min(scan, ws.max_row) + 1):
        for c in range(1, ws.max_column + 1):
            if "NOMBRE" in _norm(ws.cell(row=r, column=c).value):
                return r, c
    raise ValueError(
        "Could not find the student-name column (no header containing 'Nombre' "
        f"in the first {scan} rows). Is this the right sheet?"
    )


def parse_sheet(path: str, sheet_name: str) -> ParsedSheet:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        # Case-insensitive sheet lookup.
        match = next((s for s in wb.sheetnames if _norm(s) == _norm(sheet_name)), None)
        if match is None:
            raise ValueError(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
        ws = wb[match]

        header_row, name_col = _find_header_row(ws)
        name_header = str(ws.cell(row=header_row, column=name_col).value).strip()

        # Grade columns = other non-empty headers, skipping '#' and auto 'ColumnaN'.
        grade_cols: list[tuple[int, str]] = []
        for c in range(1, ws.max_column + 1):
            if c == name_col:
                continue
            raw = ws.cell(row=header_row, column=c).value
            header = "" if raw is None else str(raw).strip()
            if not header or header == "#":
                continue
            if re.fullmatch(r"columna\s*\d+", header, flags=re.I):
                continue
            grade_cols.append((c, header))

        students: list[Student] = []
        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(row=r, column=name_col).value
            if name is None or not str(name).strip():
                continue
            cells = {h: ws.cell(row=r, column=c).value for c, h in grade_cols}
            students.append(Student(raw_name=str(name).strip(), cells=cells))

        return ParsedSheet(
            sheet=match,
            header_row=header_row,
            name_header=name_header,
            grade_headers=[h for _, h in grade_cols],
            students=students,
        )
    finally:
        wb.close()


def round_tenth(v: float) -> float:
    """Round to nearest 0.1, half up (4.75 -> 4.8, 4.74 -> 4.7)."""
    return math.floor(v * 10 + 0.5) / 10


def _decode_date_grade(d: datetime.date) -> float | None:
    """A grade like "4.5" that Excel stored as a date decodes as day + month/10.

    Only day 1-5 and month 1-9 can come from a 0-5 grade with one decimal, so
    anything outside that is treated as a real (non-grade) date.
    """
    if 1 <= d.day <= 5 and 1 <= d.month <= 9:
        return d.day + d.month / 10
    return None


def to_grade(raw: Any) -> tuple[str, Any]:
    """Convert a raw cell to ('skip', None) | ('ok', float) | ('bad', reason)."""
    if raw is None:
        return ("skip", None)
    if isinstance(raw, bool):  # bool is an int subclass — reject explicitly
        return ("bad", f"boolean {raw!r}")
    if isinstance(raw, (datetime.datetime, datetime.date)):
        d = raw.date() if isinstance(raw, datetime.datetime) else raw
        grade = _decode_date_grade(d)
        if grade is None:
            return ("bad", f"date {d.isoformat()} does not encode a valid grade")
        return ("ok", grade)
    if isinstance(raw, (int, float)):
        return ("ok", float(raw))
    s = str(raw).strip()
    if s == "":
        return ("skip", None)
    if s == "-":
        return ("ok", DASH_GRADE)
    try:
        return ("ok", float(s.replace(",", ".")))
    except ValueError:
        return ("bad", f"text {s!r}")


def format_grade(v: float) -> str:
    return f"{round_tenth(v):.1f}"
